import openpyxl



class dataDriven():
    def read_data(self):
        path = "C:\\Users\\fathih\\PycharmProjects\\Python Automation\\data_driven_demo\\Book1.xlsx"
        workbook = openpyxl.load_workbook(path)

        sheet=workbook.active

        rows=sheet.max_row
        cols=sheet.max_column

        print(rows)
        print(cols)

        for r in range(1,rows+1):
            for c in range(1,cols+1):
                print(sheet.cell(row=r,column=c).value,end="     ")

            print()


aa=dataDriven()
aa.read_data()